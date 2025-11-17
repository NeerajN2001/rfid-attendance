import asyncio
import websockets
import json
import threading
import sys
import time
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta, date, time as dt_time

# --- Client Configuration ---
SERVER_URL = "ws://localhost:8765"
CLIENT_NAME = "db_client" # This client's name
TARGET_NAME = "esp_client" # The name of the client this program will send messages to
# ----------------------------

# --- File Paths ---
USER_DB_FILE = "userDatabase.xlsx"
USER_LOG_FILE = "userLog.xlsx"
SETTINGS_FILE = "Settings.json"

# --- Global Data Storage (In-Memory Caches) ---
# USER_DATABASE: {user_id: {'name': user_name, 'title': user_title}}
SETTINGS = {}
USER_DATABASE = {} 

# =============================================================
# FILE MANAGEMENT FUNCTIONS
# =============================================================

def initialize_files():
    """Initializes Excel and JSON files if they do not exist."""
    print("[INIT] Checking/creating necessary files...")
    
    # 1. userDatabase.xlsx
    if not os.path.exists(USER_DB_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(['user_id', 'user_name', 'user_title'])
        
        # --- Add Default Admin User ---
        ws.append(['426E3302', 'master_admin', 'admin'])
        
        wb.save(USER_DB_FILE)
        print(f"[INIT] Created {USER_DB_FILE} with headers and default admin (426E3302).")

    # 2. userLog.xlsx
    if not os.path.exists(USER_LOG_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Log"
        ws.append(['user_id', 'tap_in_time', 'tap_out_time', 'status', 'duration'])
        wb.save(USER_LOG_FILE)
        print(f"[INIT] Created {USER_LOG_FILE} with headers.")

    # 3. Settings.json
    if not os.path.exists(SETTINGS_FILE):
        default_settings = {"reset_time": "05:00:00"}
        with open(SETTINGS_FILE, 'w') as f:
            json.dump(default_settings, f, indent=4) 
        print(f"[INIT] Created {SETTINGS_FILE} with default reset time (05:00:00).")

def save_settings():
    """Saves the current SETTINGS dictionary back to Settings.json."""
    try:
        with open(SETTINGS_FILE, 'w') as f:
            json.dump(SETTINGS, f, indent=4)
        print(f"[SAVE] Settings saved: {SETTINGS}")
    except Exception as e:
        print(f"[ERROR] Could not save settings: {e}")

def load_settings():
    """Loads settings from Settings.json."""
    global SETTINGS
    try:
        with open(SETTINGS_FILE, 'r') as f:
            SETTINGS = json.load(f)
        print(f"[LOAD] Settings loaded: {SETTINGS}")
    except Exception as e:
        print(f"[ERROR] Could not load settings: {e}. Using default.")
        SETTINGS = {"reset_time": "05:00:00"}

def save_user_db():
    """Overwrites userDatabase.xlsx using the current USER_DATABASE cache."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(['user_id', 'user_name', 'user_title'])
        
        for user_id, user_data in USER_DATABASE.items():
            ws.append([user_id, user_data['name'], user_data['title']])
        
        wb.save(USER_DB_FILE)
        print(f"[SAVE] User database saved. Total users: {len(USER_DATABASE)}")
    except Exception as e:
        print(f"[ERROR] Could not save user database: {e}")

def load_user_db():
    """Loads user data from userDatabase.xlsx into USER_DATABASE cache."""
    global USER_DATABASE
    USER_DATABASE = {}
    try:
        wb = load_workbook(USER_DB_FILE)
        ws = wb.active
        
        # Skip header row (index 1)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]: # Check if user_id exists
                # Store name and title in the cache
                USER_DATABASE[str(row[0])] = {'name': str(row[1]), 'title': str(row[2])}
        print(f"[LOAD] User database loaded. Total users: {len(USER_DATABASE)}")
    except Exception as e:
        print(f"[ERROR] Could not load user database: {e}")

def get_current_log_sheet():
    """Returns the active worksheet for the user log."""
    return load_workbook(USER_LOG_FILE).active

def save_user_log(data_to_write):
    """Appends new data to the user log file."""
    try:
        wb = load_workbook(USER_LOG_FILE)
        ws = wb.active
        ws.append(data_to_write)
        wb.save(USER_LOG_FILE)
        print(f"[SAVE] Log saved: {data_to_write}")
    except Exception as e:
        print(f"[ERROR] Could not save user log: {e}")

# =============================================================
# TIME AND DURATION HELPERS (Used by SCAN)
# =============================================================

def get_current_reset_window():
    """Calculates the start and end of the current 24-hour attendance window."""
    
    reset_time_str = SETTINGS.get("reset_time", "05:00:00")
    try:
        # Pad time if needed (e.g., '05:00' to '05:00:00')
        parts = reset_time_str.split(':')
        if len(parts) == 2:
             reset_time_str += ':00'
        h, m, s = map(int, reset_time_str.split(':'))
        reset_time = dt_time(h, m, s)
    except ValueError:
        print(f"[ERROR] Invalid reset time format: {reset_time_str}. Defaulting to 05:00:00.")
        reset_time_str = "05:00:00"
        reset_time = dt_time(5, 0, 0)

    now = datetime.now()
    today_reset = datetime.combine(date.today(), reset_time)

    if now >= today_reset:
        window_start = today_reset
        window_end = today_reset + timedelta(days=1)
    else:
        window_start = today_reset - timedelta(days=1)
        window_end = today_reset
        
    return window_start, window_end, reset_time_str

def find_last_entry(user_id, ws, window_start, window_end):
    """
    Finds the last IN or OUT entry for a user within the current 24-hour window.
    Returns: (row_index, tap_in_time, status) or (None, None, None)
    """
    last_entry = None
    last_row_index = None

    for i in range(ws.max_row, 1, -1):
        row = [cell.value for cell in ws[i]]
        log_user_id = str(row[0]) if row[0] is not None else None
        
        if log_user_id == user_id:
            try:
                tap_in_time = row[1]
                log_datetime = None
                if isinstance(tap_in_time, datetime):
                    log_datetime = tap_in_time
                elif isinstance(tap_in_time, dt_time):
                     log_datetime = datetime.combine(date.today(), tap_in_time)

                if log_datetime and window_start <= log_datetime < window_end:
                    last_entry = row
                    last_row_index = i
                    break
            except Exception as e:
                print(f"[ERROR] Failed to parse datetime in log row {i}: {e}")
                continue

    if last_entry:
        return last_row_index, last_entry[1], str(last_entry[3])
    return None, None, None

def calculate_duration(start_time, end_time):
    """Calculates duration between two datetime objects and formats as HH:MM:SS."""
    if isinstance(start_time, dt_time):
        start_dt = datetime.combine(date.today(), start_time)
        end_dt = datetime.combine(date.today(), end_time)
    else:
        start_dt = start_time
        end_dt = end_time

    duration = end_dt - start_dt
    
    total_seconds = int(duration.total_seconds())
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

# =============================================================
# COMMAND HANDLERS
# =============================================================

async def send_reply(websocket, reply_msg):
    """Utility function to format and send a reply back to the ESP client."""
    msg = {"to": TARGET_NAME, "msg": reply_msg}
    await websocket.send(json.dumps(msg))
    print(f"[REPLY] ⬅️ Sent to {TARGET_NAME}: {reply_msg}")

async def handle_scan_command(websocket, id_value):
    """Handles the SCAN command (Attendance Logic)."""
    
    user_data = USER_DATABASE.get(id_value)
    user_name = user_data['name'] if user_data else None
    
    # Check if the user is an admin
    is_admin = user_data and user_data.get('title', '').lower() == 'admin'
    
    now = datetime.now()
    current_time_str = now.strftime('%H:%M:%S')
    ws = get_current_log_sheet()
    window_start, window_end, reset_time_str = get_current_reset_window()

    # --- Step 1: User Not Found ---
    if not user_name:
        reply = {"md": "scan", "rslt": "NF"}
        return await send_reply(websocket, reply)

    # --- Find Last Log Entry within current 24h window ---
    last_row_index, last_tap_in_time, last_status = find_last_entry(id_value, ws, window_start, window_end)

    # --- Core Logic Flow (Prioritized) ---
    
    # 4. ADMIN EXEMPTION: If the admin tapped OUT, reset the status to None (ready for Tap IN)
    if is_admin and last_status == 'Tap Out':
        print("[LOGIC] Admin user reset their status to Tap IN available (exempted from one-entry rule).")
        last_status = None 
    
    # 3. Completed Cycle (FAIL for regular users)
    if last_status == 'Tap Out':
        # This condition is only reached by non-admin users
        reply = {"md": "scan", "nm": user_name, "act": "FAIL", "rtr": reset_time_str}
        return await send_reply(websocket, reply)

    # 2. Tap In Exists (Tap OUT for all)
    elif last_status == 'Tap In':
        
        duration_str = calculate_duration(last_tap_in_time, now)
        
        try:
            wb = load_workbook(USER_LOG_FILE)
            ws = wb.active
            
            # Note: last_row_index cannot be None here
            ws.cell(row=last_row_index, column=3, value=now)
            ws.cell(row=last_row_index, column=4, value='Tap Out')
            ws.cell(row=last_row_index, column=5, value=duration_str)
            wb.save(USER_LOG_FILE)
            
            print(f"[SAVE] Log updated (Tap Out). Duration: {duration_str}")

            reply = {"md": "scan", "nm": user_name, "act": "OUT", "tm": current_time_str, "dur": duration_str}
            return await send_reply(websocket, reply)
        except Exception as e:
            print(f"[ERROR] Failed to update user log file for Tap Out: {e}")
            reply = {"md": "scan", "rslt": "LOG_ERR"}
            return await send_reply(websocket, reply)

    # 1. No Entry Found (Tap IN for all, including admins after Tap Out)
    elif last_status is None:
        
        new_entry = [id_value, now, None, 'Tap In', None]
        save_user_log(new_entry)
        
        reply = {"md": "scan", "nm": user_name, "act": "IN", "tm": current_time_str}
        return await send_reply(websocket, reply)
        
    else:
        reply = {"md": "scan", "rslt": "UNKNOWN_ERR"}
        return await send_reply(websocket, reply)

async def handle_auth_command(websocket, id_value):
    """Handles the AUTH command (Admin check)."""
    user_data = USER_DATABASE.get(id_value)
    
    if user_data and user_data.get('title', '').lower() == 'admin':
        reply = {"md": "auth", "rslt": "admin"}
    else:
        reply = {"md": "auth", "rslt": "not-admin"}
        
    return await send_reply(websocket, reply)

async def handle_search_command(websocket, id_value):
    """Handles the SEARCH command (User existence check)."""
    
    if id_value in USER_DATABASE:
        reply = {"md": "search", "rslt": "F"} # Found
    else:
        reply = {"md": "search", "rslt": "NF"} # Not Found
        
    return await send_reply(websocket, reply)

async def handle_add_command(websocket, id_value, username, usertitle):
    """Handles the ADD command (Add new user)."""
    
    # Update in-memory cache
    USER_DATABASE[id_value] = {'name': username, 'title': usertitle}
    
    # Save the updated database to the Excel file
    save_user_db()
    
    reply = {"md": "add", "rslt": "A"}
    return await send_reply(websocket, reply)

async def handle_delete_command(websocket, id_value):
    """Handles the DELETE command (Delete existing user)."""
    
    if id_value in USER_DATABASE:
        # Delete from in-memory cache
        del USER_DATABASE[id_value]
        
        # Save the updated database to the Excel file
        save_user_db()
        
        reply = {"md": "delete", "rslt": "DL"}
    else:
        # Reply DL even if not found, as the effect is the same (user is gone)
        reply = {"md": "delete", "rslt": "DL"} 

    return await send_reply(websocket, reply)

async def handle_reset_time_command(websocket, time_value):
    """Handles the RST_TIME command (Update reset time)."""
    
    # Basic validation (simple HH:MM:SS format check)
    if not (2 <= time_value.count(':') <= 3):
        print(f"[ERROR] Invalid time format received: {time_value}")
        reply = {"md": "rst_time", "rslt": "FAIL"}
        return await send_reply(websocket, reply)
        
    # Update global settings and save
    SETTINGS['reset_time'] = time_value
    save_settings()
    
    # Reload the user database to ensure integrity (optional, but good practice)
    load_user_db()
    
    reply = {"md": "rst_time", "rslt": "D"}
    return await send_reply(websocket, reply)

# =============================================================
# ASYNC CLIENT FUNCTIONS
# =============================================================

async def receive_loop(websocket):
    """Asynchronously listens for and processes messages."""
    try:
        async for message in websocket:
            try:
                data = json.loads(message)
                
                # Check for forwarded messages from the ESP client
                if data.get("from") == TARGET_NAME and "msg" in data:
                    payload = data["msg"]
                    command_md = payload.get("md")
                    
                    print(f"\n[RECEIVED CMD] '{command_md}' from {TARGET_NAME}")
                    
                    if command_md == "scan" and "id" in payload:
                        await handle_scan_command(websocket, payload["id"])
                        
                    elif command_md == "auth" and "id" in payload:
                        await handle_auth_command(websocket, payload["id"])

                    elif command_md == "search" and "id" in payload:
                        await handle_search_command(websocket, payload["id"])

                    elif command_md == "add" and all(k in payload for k in ["id", "un", "ut"]):
                        await handle_add_command(websocket, payload["id"], payload["un"], payload["ut"])

                    elif command_md == "delete" and "id" in payload:
                        await handle_delete_command(websocket, payload["id"])

                    elif command_md == "rst_time" and "tm" in payload:
                        await handle_reset_time_command(websocket, payload["tm"])

                    else:
                        print(f"\n[RECEIVED from {TARGET_NAME}]: Unknown/Malformed command: {payload}")
                
                # Handle server status/error messages
                elif "status" in data or "error" in data:
                    print(f"\n[SERVER STATUS/ERROR]: {data.get('message') or data.get('error')}")
                
                else:
                    print(f"\n[RAW MESSAGE] (Not from {TARGET_NAME}): {message}")

            except json.JSONDecodeError:
                print(f"\n[ERROR] Received invalid JSON: {message}")
            except Exception as e:
                print(f"\n[ERROR] An error occurred in receiver: {e}")
                break

    except websockets.exceptions.ConnectionClosed:
        print("\n[INFO] Connection closed by the server.")
    finally:
        print("[INFO] Receiver loop stopping.")
        os._exit(0) 

async def connect_and_run():
    """Establishes connection and manages sender/receiver coroutines."""
    # 1. Initialize and load files before connecting
    initialize_files()
    load_settings()
    load_user_db()
    
    try:
        # Note: Using localhost requires the server to be running on the same machine
        async with websockets.connect(SERVER_URL) as websocket:
            
            # Send registration message immediately
            registration_msg = json.dumps({"type": "register", "name": CLIENT_NAME})
            await websocket.send(registration_msg)
            
            # Run the receiver loop
            await receive_loop(websocket)

    except ConnectionRefusedError:
        print(f"\n[FATAL] Connection refused. Is the server running at {SERVER_URL}?")
    except Exception as e:
        print(f"\n[FATAL] An unexpected connection error occurred: {e}")
    finally:
        await asyncio.sleep(0.5)
        print("[INFO] Client application closing.")

if __name__ == "__main__":
    import os
    try:
        # The client now runs a single async function that loops forever
        asyncio.run(connect_and_run())
    except KeyboardInterrupt:
        print("\n[INFO] Client application interrupted.")
    except Exception as e:
        print(f"\n[FATAL] Main loop error: {e}")